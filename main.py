import cv2
import math
import pandas as pd
import tkinter as tk
import cvzone
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from PIL import Image, ImageTk
from ultralytics import YOLO
from tracker import*

# model và video
model=YOLO('best.pt')
cap=cv2.VideoCapture(r"C:\Users\Acer\OneDrive\Documents\TDH\university\năm 3\học kỳ 1\xử lý ảnh\De9_Duong_Xuan_Ngoc_211602765\code\n1.mp4")

# thời gian bắt đầu chạy program
previous_time = datetime.now().date()

# tạo một file excel
workbook = Workbook()
sheet = workbook.active
sheet.append(["Thoi gian", "Xe dap", "Xe may", "O to", "Tổng"])
for cell in sheet["1:1"]:
    cell.font = Font(bold=True)
sheet.cell(row=2, column=1, value=previous_time)
sheet.cell(row=2, column=2, value=0)
sheet.cell(row=2, column=3, value=0)
sheet.cell(row=2, column=4, value=0)
sheet.cell(row=2, column=5, value=0)

# phân loại tên xe 
class_list = ["bicycle", "motorcycle", "car"]

# tạo giao diện 

# các biến
count=0
cy1=424
offset=15
current_row = 2
counter1=[]
counter2=[]
counter3=[]

# thuật toán theo dõi
tracker1=Tracker()
tracker2=Tracker()
tracker3=Tracker()

# tạo giao diện
window = tk.Tk()
window.title("Show")
window.geometry("1050x610")
large_font = ('Times New Roman', 16)
label = tk.Label(window)
label.place(x = 10, y = 10)
label1 = tk.Label(window, font = large_font)
label1.place(x=10, y=560)
label2 = tk.Label(window, text = "Time", font = large_font)
label2.place(x = 10, y = 520)
label3 = tk.Label(window, text = "Bicycle", font = large_font)
label3.place(x = 140, y = 520)
label4 = tk.Label(window, text = "Motorcycle", font = large_font)
label4.place(x = 220, y = 520)
label5 = tk.Label(window, text = "Car", font = large_font)
label5.place(x = 330, y = 520)
label6 = tk.Label(window, text = "Total", font = large_font)
label6.place(x = 370, y = 520)
label7 = tk.Label(window, font = large_font)
label7.place(x=145, y=560)
label8 = tk.Label(window, font = large_font)
label8.place(x=225, y=560)
label9 = tk.Label(window, font = large_font)
label9.place(x=335, y=560)
label10 = tk.Label(window, font = large_font)
label10.place(x=375, y=560)

while True:    
    ret,frame = cap.read()
    if not ret:
        break

    count += 1
    if count % 3 != 0:
        continue
    
    frame=cv2.resize(frame,(1020,500))
   

    results=model.predict(frame)
    # sang ngày sẽ cập nhật lại dữ liệu
    current_time = datetime.now().date()
    
    label1.config(text=current_time)
    if current_time > previous_time:
        current_row += 1
    # tạo dữ liệu cho phương tiện
    a=results[0].boxes.data
    px=pd.DataFrame(a).astype("float")

    # tạo các mảng rỗng
    list1=[]
    motorcycle=[]
    list2=[]
    car=[]
    list3=[]
    bicycle=[]
    
    # tạo line đếm xe 
    cv2.line(frame,(100,cy1),(1100,cy1),(0,0,255),2)

    # gán nhãn cho phương tiện
    for r in results:
        boxes = r.boxes
        for box in boxes:
            x1, y1, x2, y2 = box.xyxy[0]
            x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
            w, h = x2 - x1, y2 - y1
            cx, cy = x1+w//2, y1+h//2
            cv2.circle(frame,(cx,cy),5,(255,0,255),cv2.FILLED)
            conf = math.ceil((box.conf[0]*100)) / 100 
            cls = int(box.cls[0])
            currentclass = class_list[cls]
            if currentclass in ["bicycle", "motorcycle", "car"] and conf > 0.5:
                cvzone.cornerRect(frame, (x1, y1, w, h), l=15, rt=2, colorR=(255,0,0))
                cvzone.putTextRect(frame,f'{class_list[cls]} {conf}', (max(0, x1), max(35, y1)), scale=0.6, thickness = 1, offset=3)
                
    # in các thông số trong hàng của từng index
    for index,row in px.iterrows():

        x1=int(row[0])
        y1=int(row[1])
        x2=int(row[2])
        y2=int(row[3])
        d=int(row[5])
        c=class_list[d]
        if 'motorcycle' in c:
            list1.append([x1,y1,x2,y2])
            motorcycle.append(c)
        if 'car' in c:
            list2.append([x1,y1,x2,y2])
            car.append(c)
        if 'bicycle' in c:
            list3.append([x1,y1,x2,y2])
            bicycle.append(c) 

    # theo dõi các đối tượng
    bbox1_idx=tracker1.update(list1)
    bbox2_idx=tracker2.update(list2)
    bbox3_idx=tracker3.update(list3)

    # đếm xe máy
    for bbox1 in bbox1_idx:
        for i in motorcycle:
            x5,y5,x6,y6,id2=bbox1
            cxm=int(x5+x6)//2
            cym=int(y5+y6)//2
            if cym<(cy1+offset) and cym>(cy1-offset):
               cv2.circle(frame,(cxm,cym),4,(0,255,0),-1)
               
               if counter1.count(id2)==0:
                  counter1.append(id2)
                  cv2.line(frame,(100,cy1),(1100,cy1),(0,255,0),2)
                  formatted_time = current_time.strftime("%Y-%m-%d")
                  sheet.cell(row=current_row, column=1, value=formatted_time)
                  sheet.cell(row=current_row, column=4, value=len(counter1))

    # đếm xe ô tô          
    for bbox2 in bbox2_idx:
        for i in car:
            x5,y5,x6,y6,id2=bbox2
            cxm=int(x5+x6)//2
            cym=int(y5+y6)//2
            if cym<(cy1+offset) and cym>(cy1-offset):
               cv2.circle(frame,(cxm,cym),4,(0,255,0),-1)
               
               if counter2.count(id2)==0:
                  counter2.append(id2)
                  cv2.line(frame,(100,cy1),(1100,cy1),(0,255,0),2)
                  formatted_time = current_time.strftime("%Y-%m-%d")
                  sheet.cell(row=current_row, column=1, value=formatted_time)
                  sheet.cell(row=current_row, column=4, value=len(counter2))
    # đếm xe đạp 
    for bbox3 in bbox3_idx:
        for i in bicycle:
            x5,y5,x6,y6,id2=bbox3
            cxm=int(x5+x6)//2
            cym=int(y5+y6)//2
            if cym<(cy1+offset) and cym>(cy1-offset):
               cv2.line(frame,(2,cy1),(794,cy1),(0,255,0),2)
               
               cvzone.putTextRect(frame,f'{id2}',(x5,y5),1,1)
               if counter3.count(id2)==0:
                counter3.append(id2)
                cv2.line(frame,(100,cy1),(1100,cy1),(0,255,0),2)
                formatted_time = current_time.strftime("%Y-%m-%d")
                sheet.cell(row=current_row, column=1, value=formatted_time)
                sheet.cell(row=current_row, column=4, value=len(counter3))
    
    # xuất dữ liệu hiển thị lên giao diện
    motorcyclec=(len(counter1))
    carr=(len(counter2))
    bicyclee=(len(counter3))
    s = (len(counter1))+(len(counter2))+(len(counter3))
    label7.config(text=bicyclee)
    label8.config(text=motorcyclec)
    label9.config(text=carr)
    label10.config(text=s)
    photo = ImageTk.PhotoImage(Image.fromarray(frame))
    label.config(image=photo)
    label.photo = photo

    # lưu dữ liệu lên excel
    sheet.cell(row=current_row, column=5, value=s)
    workbook.save("xlad9.xlsx")


    window.update()

cap.release()
cv2.destroyAllWindows()
window.mainloop()


